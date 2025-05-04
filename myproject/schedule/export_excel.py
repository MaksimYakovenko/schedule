from schedule.models import (ScheduleEntry)
from collections import defaultdict
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl import Workbook



def export_schedule_excel(request):
    wb = Workbook()
    bold_font = Font(bold=True)

    default_sheet = wb.active
    wb.remove(default_sheet)

    entries = ScheduleEntry.objects.select_related(
        'lesson__teacher', 'lesson__subject', 'group', 'classroom'
    ).order_by('lesson__course', 'group__name', 'day_of_week', 'lesson_number')

    schedule_by_course = defaultdict(list)
    for entry in entries:
        course = entry.lesson.course
        schedule_by_course[course].append(entry)

    for course, course_entries in schedule_by_course.items():
        sheet_name = f"{course} курс"
        ws = wb.create_sheet(title=sheet_name[:31])

        headers = ["Група", "День", "Номер пари", "Тип", "Предмет", "Викладач", "Аудиторія", "Період"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header).font = bold_font

        for row_num, entry in enumerate(course_entries, start=2):
            ws.cell(row=row_num, column=1, value=entry.group.name)
            ws.cell(row=row_num, column=2, value=entry.day_of_week)
            ws.cell(row=row_num, column=3, value=entry.lesson_number)
            ws.cell(row=row_num, column=4, value=entry.lesson.lesson_type)
            ws.cell(row=row_num, column=5, value=entry.lesson.subject.name)
            ws.cell(row=row_num, column=6, value=entry.lesson.teacher.full_name)
            ws.cell(row=row_num, column=7, value=entry.classroom.name)

            if entry.lesson.start_date and entry.lesson.end_date:
                period = f"з {entry.lesson.start_date.strftime('%d.%m.%Y')} до {entry.lesson.end_date.strftime('%d.%m.%Y')}"
            else:
                period = ""
            ws.cell(row=row_num, column=8, value=period)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=rozklad_po_kursam.xlsx"
    wb.save(response)
    return response