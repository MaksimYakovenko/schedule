from django.shortcuts import render, get_object_or_404
import random
from django.urls import reverse_lazy, reverse
from django.views.generic import CreateView, ListView, DeleteView, UpdateView

from schedule.models import (Department, Teacher, Classroom, Lesson, Group,
                             Subject, Semester, ScheduleEntry)
from django.shortcuts import redirect
from collections import defaultdict
from django.views.decorators.http import require_POST
from .forms import ScheduleEntryForm, EditScheduleEntryForm
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.utils.html import strip_tags
from openpyxl.styles import Font
from openpyxl import Workbook



def home(request):
    teacher_count = Teacher.objects.count()
    department_count = Department.objects.count()
    lessons_count = ScheduleEntry.objects.count()
    return render(request, 'home.html', {
        'teacher_count': teacher_count,
        'department_count': department_count,
        'lessons_count': lessons_count
    })



DAYS_OF_WEEK = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', 'Пʼятниця']
LESSON_NUMBERS = [1, 2, 3, 4]
LESSON_TIMES = {
    1: '08:00-09:00',
    2: '10:00-11:00',
    3: '12:00-13:00',
    4: '14:00-15:00',
}


def generate_schedule(request):
    if request.method == 'POST':
        semester_id = request.POST.get('semester_id')
        if not semester_id:
            return HttpResponseBadRequest("Семестр не вибрано.")

    ScheduleEntry.objects.all().delete()

    groups = Group.objects.filter(lesson__isnull=False).distinct()
    classrooms = Classroom.objects.all()

    teacher_occupancy = defaultdict(set)
    classroom_occupancy = defaultdict(set)
    group_occupancy = defaultdict(set)
    schedule_dict = {}

    for group in groups:
        group_lessons = Lesson.objects.filter(group=group)

        for lesson in group_lessons:
            lessons_scheduled = 0
            attempts = 0
            max_attempts = 100

            while lessons_scheduled < lesson.hours_per_week and attempts < max_attempts:
                attempts += 1
                day = random.choice(DAYS_OF_WEEK)
                lesson_number = random.choice(LESSON_NUMBERS)

                if (
                    lesson.teacher.id in teacher_occupancy[(day, lesson_number)]
                    or group.id in group_occupancy[(day, lesson_number)]
                ):
                    continue

                available_classrooms = [
                    classroom for classroom in classrooms
                    if classroom.id not in classroom_occupancy[(day, lesson_number)]
                ]

                if not available_classrooms:
                    continue

                classroom = random.choice(available_classrooms)

                ScheduleEntry.objects.create(
                    group=group,
                    day_of_week=day,
                    lesson_number=lesson_number,
                    lesson=lesson,
                    classroom=classroom
                )

                teacher_occupancy[(day, lesson_number)].add(lesson.teacher.id)
                classroom_occupancy[(day, lesson_number)].add(classroom.id)
                group_occupancy[(day, lesson_number)].add(group.id)

                if group.id not in schedule_dict:
                    schedule_dict[group.id] = {}
                schedule_dict[group.id].setdefault(day, []).append(lesson_number)

                lessons_scheduled += 1

    request.session['schedule_dict'] = schedule_dict
    return redirect(f"{reverse('schedule_view')}?semester_id={semester_id}")

def schedule_view(request):
    groups = Group.objects.all()
    teachers = Teacher.objects.all()
    semesters = Semester.objects.all()
    teacher_id = request.GET.get('teacher_id')
    semester_id = request.GET.get('semester_id')
    course = request.GET.get('course')
    group_id = request.GET.get('group_id')

    schedule_entries = ScheduleEntry.objects.all()

    if semester_id in [None, '', 'None']:
        semester_id = None

    if teacher_id in [None, '', 'None']:
        teacher_id = None

    if group_id in [None, '', 'None']:
        group_id = None

    if teacher_id:
        schedule_entries = schedule_entries.filter(
            lesson__teacher_id=teacher_id)

    if semester_id:
        schedule_entries = schedule_entries.filter(
            lesson__semester_id=semester_id)

    if group_id:
        schedule_entries = schedule_entries.filter(
            lesson__group_id=group_id)

    if course:
        schedule_entries = schedule_entries.filter(lesson__course=course)


    schedule_entries = sorted(
        schedule_entries,
        key=lambda x: (DAYS_OF_WEEK.index(x.day_of_week), x.lesson_number)
    )

    groups_with_schedule = []

    for group in groups:
        has_lessons = any(entry.group == group for entry in schedule_entries)
        if has_lessons:
            groups_with_schedule.append(group)

    schedule_availability = {}
    for group in groups:
        schedule_availability[group.id] = {}
        for day in DAYS_OF_WEEK:
            schedule_availability[group.id][day] = False

        for entry in schedule_entries:
            if entry.group.id not in schedule_availability:
                schedule_availability[entry.group.id] = {}

            schedule_availability[entry.group.id][entry.day_of_week] = True

    groups_by_course = defaultdict(list)
    for group in groups_with_schedule:
        lessons = Lesson.objects.filter(group=group)
        distinct_courses = lessons.values_list('course', flat=True).distinct()

        for course in distinct_courses:
            groups_by_course[course].append(group)

    sorted_course_keys = sorted(
        groups_by_course.keys(),
        key=lambda x: (int(x[1]) if x.startswith('m') else int(x))
        if x.isdigit() else (5 if x == 'm1' else 6)
    )

    context = {
        'groups_by_course': dict(groups_by_course),
        'sorted_course_keys': sorted_course_keys,
        'groups_with_schedule': groups_with_schedule,
        'days_of_week': DAYS_OF_WEEK,
        'lesson_numbers': LESSON_NUMBERS,
        'groups': groups,
        'schedule_entries': schedule_entries,
        'lesson_times': LESSON_TIMES,
        'teachers': teachers,
        'selected_teacher_id': teacher_id,
        'selected_semester_id': semester_id,
        'selected_group_id': group_id,
        'semesters': semesters,
        'schedule_availability': schedule_availability,
        'selected_course': course,
    }
    return render(request, 'schedule.html', context)


@require_POST
def delete_lesson(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)
    lesson.delete()
    return redirect('schedule_view')


def edit_lesson(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)

    if request.method == 'POST':
        form = LessonForm(request.POST, instance=lesson)
        if form.is_valid():
            form.save()
            return redirect('schedule_view')
    else:
        form = LessonForm(instance=lesson)

    return render(request, 'schedule/edit_lesson.html', {'form': form, 'lesson': lesson})


def add_lesson_view(request):
    if request.method == 'POST':
        form = ScheduleEntryForm(request.POST)
        if form.is_valid():
            course = form.cleaned_data.get('course')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')

            teacher = form.cleaned_data['teacher']
            classroom = form.cleaned_data['classroom']
            day_of_week = form.cleaned_data['day_of_week']
            lesson_number = form.cleaned_data['lesson_number']
            group = form.cleaned_data['group']
            semester = form.cleaned_data['semester']

            teacher_conflict = ScheduleEntry.objects.filter(
                lesson__teacher=teacher,
                day_of_week=day_of_week,
                lesson_number=lesson_number,
                lesson__semester=semester,
            ).exists()

            if teacher_conflict:
                form.add_error('teacher',
                               'Викладач зайнятий у цей день і пару.')

            classroom_conflict = ScheduleEntry.objects.filter(
                classroom=classroom,
                day_of_week=day_of_week,
                lesson_number=lesson_number,
                lesson__semester=semester,
            ).exists()

            if classroom_conflict:
                form.add_error('classroom',
                               'Аудиторія зайнята у цей день і пару.')

            group_conflict = ScheduleEntry.objects.filter(
                group=group,
                day_of_week=day_of_week,
                lesson_number=lesson_number,
                lesson__semester=semester,
            ).exists()

            if group_conflict:
                form.add_error('group', 'У групи вже є пара в цей день і пару.')

            if not form.errors and course not in [None, '']:
                lesson, _ = Lesson.objects.get_or_create(
                    subject=form.cleaned_data['subject'],
                    teacher=teacher,
                    course=course,
                    start_date=start_date,
                    end_date=end_date,
                    lesson_type=form.cleaned_data['lesson_type'],
                    group=group,
                    semester=semester,
                    classroom=classroom,
                )
                ScheduleEntry.objects.get_or_create(
                    lesson=lesson,
                    group=group,
                    classroom=classroom,
                    day_of_week=day_of_week,
                    lesson_number=lesson_number,
                )
                return redirect('schedule_view')

            elif course in [None, '']:
                form.add_error('course', 'Поле "Курс" є обов’язковим.')

    else:
        form = ScheduleEntryForm()

    return render(request, 'schedule/add_lesson_entry.html', {'form': form})



def export_schedule_excel(request):
    wb = Workbook()
    bold_font = Font(bold=True)

    default_sheet = wb.active
    wb.remove(default_sheet)

    entries = ScheduleEntry.objects.select_related(
        'lesson__teacher', 'lesson__subject', 'group', 'classroom'
    ).order_by('lesson__course', 'group__name', 'day_of_week', 'lesson_number')

    schedule_by_course = defaultdict(list)
    for entry in entries:
        course = entry.lesson.course
        schedule_by_course[course].append(entry)

    for course, course_entries in schedule_by_course.items():
        sheet_name = f"{course} курс"
        ws = wb.create_sheet(title=sheet_name[:31])

        headers = ["Група", "День", "Номер пари", "Тип", "Предмет", "Викладач", "Аудиторія", "Період"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header).font = bold_font

        for row_num, entry in enumerate(course_entries, start=2):
            ws.cell(row=row_num, column=1, value=entry.group.name)
            ws.cell(row=row_num, column=2, value=entry.day_of_week)
            ws.cell(row=row_num, column=3, value=entry.lesson_number)
            ws.cell(row=row_num, column=4, value=entry.lesson.lesson_type)
            ws.cell(row=row_num, column=5, value=entry.lesson.subject.name)
            ws.cell(row=row_num, column=6, value=entry.lesson.teacher.full_name)
            ws.cell(row=row_num, column=7, value=entry.classroom.name)

            if entry.lesson.start_date and entry.lesson.end_date:
                period = f"з {entry.lesson.start_date.strftime('%d.%m.%Y')} до {entry.lesson.end_date.strftime('%d.%m.%Y')}"
            else:
                period = ""
            ws.cell(row=row_num, column=8, value=period)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=rozklad_po_kursam.xlsx"
    wb.save(response)
    return response


class AddDepartment(CreateView):
    model = Department
    template_name = 'add_department.html'
    success_url = reverse_lazy('department_list')
    fields = '__all__'


class AddSubject(CreateView):
    model = Subject
    template_name = 'add_subject.html'
    success_url = reverse_lazy('subject_list')
    fields = '__all__'


class AddSemester(CreateView):
    model = Semester
    template_name = 'add_semester.html'
    success_url = reverse_lazy('semester_list')
    fields = '__all__'


class SemesterListView(ListView):
    model = Semester
    template_name = 'semester_list.html'
    context_object_name = 'semesters'


class LessonListView(ListView):
    model = Lesson
    template_name = 'lesson_list.html'
    context_object_name = 'lessons'


class SemesterDeleteView(DeleteView):
    model = Semester
    template_name = 'semester_confirm_delete.html'
    success_url = reverse_lazy('semester_list')
    fields = '__all__'


class SemesterUpdateView(UpdateView):
    model = Semester
    template_name = 'edit_semester.html'
    fields = '__all__'
    success_url = reverse_lazy('semester_list')


class SubjectDeleteView(DeleteView):
    model = Subject
    template_name = 'add_subject.html'
    success_url = reverse_lazy('subject_list')
    fields = '__all__'


class SubjectUpdateView(UpdateView):
    model = Subject
    template_name = 'edit_subject.html'
    fields = '__all__'
    success_url = reverse_lazy('subject_list')


class DepartmentListView(ListView):
    model = Department
    template_name = 'department_list.html'
    context_object_name = 'departments'


class SubjectListView(ListView):
    model = Subject
    template_name = 'subject_list.html'
    context_object_name = 'subjects'


class GroupListView(ListView):
    model = Group
    template_name = 'group_list.html'
    context_object_name = 'groups'


class AudienceListView(ListView):
    model = Classroom
    template_name = 'audience_list.html'
    context_object_name = 'audiences'


class DepartmentDeleteView(DeleteView):
    model = Department
    template_name = 'department_confirm_delete.html'
    success_url = reverse_lazy('department_list')


class AddTeacher(CreateView):
    model = Teacher
    template_name = 'add_teacher.html'
    success_url = reverse_lazy('teacher_list')
    fields = '__all__'


class AddLesson(CreateView):
    model = Lesson
    template_name = 'add_lesson.html'
    success_url = reverse_lazy('lesson_list')
    fields = '__all__'


class AddGroup(CreateView):
    model = Group
    template_name = 'add_group.html'
    success_url = reverse_lazy('group_list')
    fields = '__all__'


class LessonDeleteView(DeleteView):
    model = Lesson
    template_name = 'lesson_confirm_delete.html'
    success_url = reverse_lazy('lesson_list')



class GroupDeleteView(DeleteView):
    model = Group
    template_name = 'group_confirm_delete.html'
    success_url = reverse_lazy('group_list')


class GroupUpdateView(UpdateView):
    model = Group
    template_name = 'edit_group.html'
    fields = '__all__'
    success_url = reverse_lazy('group_list')


class LessonUpdateView(UpdateView):
    model = Lesson
    template_name = 'edit_lesson.html'
    fields = '__all__'
    success_url = reverse_lazy('lesson_list')


class AddAudience(CreateView):
    model = Classroom
    template_name = 'add_audience.html'
    success_url = reverse_lazy('audience_list')
    fields = '__all__'


class TeacherListView(ListView):
    model = Teacher
    template_name = 'teacher_list.html'
    context_object_name = 'teachers'

    def get_queryset(self):
        queryset = super().get_queryset()
        department = self.request.GET.get('department', None)
        if department:
            queryset = queryset.filter(department__name=department)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departments'] = Department.objects.values_list('name',
                                                                flat=True).distinct()
        context['selected_department'] = self.request.GET.get('department', '')
        return context


class TeacherDeleteView(DeleteView):
    model = Teacher
    template_name = 'teacher_confirm_delete.html'
    success_url = reverse_lazy('teacher_list')


class DepartmentUpdateView(UpdateView):
    model = Department
    template_name = 'edit_department.html'
    fields = '__all__'
    success_url = reverse_lazy('department_list')


class TeacherUpdateView(UpdateView):
    model = Teacher
    template_name = 'edit_teacher.html'
    fields = '__all__'
    success_url = reverse_lazy('teacher_list')


class AudienceDeleteView(DeleteView):
    model = Classroom
    template_name = 'audience_confirm_delete.html'
    success_url = reverse_lazy('audience_list')


class AudienceUpdateView(UpdateView):
    model = Classroom
    template_name = 'edit_audience.html'
    fields = '__all__'
    success_url = reverse_lazy('audience_list')


class About(CreateView):
    model = Teacher
    template_name = 'about.html'
    fields = '__all__'


class Contacts(CreateView):
    model = Teacher
    template_name = 'contact.html'
    fields = '__all__'
